#!/usr/bin/env python3
"""
匯入 Excel 課程資料到 Oracle 資料庫
使用 openpyxl 或將 Excel 轉為 CSV 後處理
包含完整流程：
1. 檢查並建立缺失的主題和細項主題
2. 建立學年期主題設定和細項主題設定
3. 匯入課程填寫記錄
"""

import csv
import asyncio
import sys
import re
from typing import Dict, List, Optional, Set, Tuple
from datetime import datetime
import tempfile
import os

import oracledb
from course_selection_api.config import get_settings
from course_selection_api.data_access_object.db import Database, get_database_dsn
from course_selection_api.data_access_object.theme_dao import ThemeDAO, SubThemeDAO
from course_selection_api.data_access_object.school_year_settings_dao import (
    SchoolYearThemeSettingsDAO,
    SchoolYearSubThemeSettingsDAO,
    CourseEntriesDAO
)

# 設定
settings = get_settings()

# 嘗試導入 openpyxl，如果沒有則使用 csv 方法
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: 未安裝 openpyxl，將嘗試使用其他方法")

# Excel 欄位到主題和細項主題的映射（動態建立，支援多種格式）
# 格式: {excel_column_name: (theme_code, sub_theme_code, week_column_name)}
# 注意: Excel 使用數字代碼格式，我們需要根據實際欄位動態建立映射

# 數字代碼到資料庫 sub_theme_code 的映射
EXCEL_CODE_TO_SUB_THEME = {
    # A301 主題代碼映射
    '2030': ('A301', '1012', None, '跨域'),  # 2030: 跨域
    '2020': ('A301', '1020', '實作週次', '實作'),  # 2020: 實作
    '2050': ('A301', '1030', None, '行動導向'),  # 2050: 行動導向
    '4010': ('A301', '1040', None, '氣候變遷'),  # 4010: 氣候變遷
    '4015': ('A301', '1050', None, '淨零排放'),  # 4015: 淨零排放
    '4020': ('A301', '1060', None, '永續發展'),  # 4020: 永續發展
    '4025': ('A301', '1070', None, '情緒管理'),  # 4025: 情緒管理
    '2010': ('A301', '1080', '人文關懷週次', '人文關懷'),  # 2010: 人文關懷
    '4030': ('A301', '1090', None, '性別平等'),  # 4030: 性別平等
    '4035': ('A301', '1100', None, '生命教育'),  # 4035: 生命教育
    '2048': ('A301', '1110', None, '資訊安全'),  # 2048: 資訊安全
    '2044': ('A301', '1120', '媒體識讀或資訊判讀週次', '資訊判讀'),  # 2044: 資訊判讀
    '2042': ('A301', '1130', '媒體識讀或資訊判讀週次', '媒體識讀'),  # 2042: 媒體識讀
    '4040': ('A301', '1140', None, '智慧財產'),  # 4040: 智慧財產
    '4045': ('A301', '1150', None, '臺灣文學'),  # 4045: 臺灣文學
    '4050': ('A301', '1160', None, '走讀台中山水遊學'),  # 4050: 走讀台中山水遊學
    '4055': ('A301', '1170', None, '空間綠化設計'),  # 4055: 空間綠化設計
    '4060': ('A301', '1180', None, '生活環境創意美學'),  # 4060: 生活環境創意美學
    '4065': ('A301', '1190', None, '運算思維程式設計'),  # 4065: 運算思維程式設計
    '4070': ('A301', '1200', None, '創新創意'),  # 4070: 創新創意/創新創業
    '4075': ('A301', '1210', None, '設計思考'),  # 4075: 設計思考
    '3010': ('A301', '1220', None, '生成式AI'),  # 3010: 生成式AI
    '2060': ('A301', '1230', None, '職涯發展'),  # 2060: 職涯發展
    
    # USR (A501) 代碼映射
    '2015': ('A501', '3010', '在地關懷週次', '在地關懷'),  # 2015: 在地關懷
    '5010': ('A501', '3020', None, '文化永續'),  # 5010: 文化永續
    '5020': ('A501', '3030', None, '健康促進'),  # 5020: 健康促進
    '5030': ('A501', '3040', None, '永續環境'),  # 5030: 永續環境
    '5040': ('A501', '3050', None, '產業鏈結'),  # 5040: 產業鏈結
    '5050': ('A501', '3060', None, '食品安全'),  # 5050: 食品安全
    '5060': ('A501', '3070', None, '社會實踐'),  # 5060: 社會實踐
    
    # UCAN (A401) 代碼映射
    '2040': ('A401', '1010', '資訊科技週次', '資訊科技'),  # 2040: 資訊科技
    '6010': ('A401', '2010', None, 'UCAN創新'),  # 6010: UCAN創新
    '6020': ('A401', '2020', None, 'UCAN問題解決'),  # 6020: UCAN問題解決
    '6030': ('A401', '2030', None, 'UCAN人際互動'),  # 6030: UCAN人際互動
    '6040': ('A401', '2040', None, 'UCAN工作責任'),  # 6040: UCAN工作責任
    '6050': ('A401', '2050', None, 'UCAN持續學習'),  # 6050: UCAN特質/持續學習
    '6060': ('A401', '2060', None, 'UCAN團隊合作'),  # 6060: UCAN團隊合作
    '6070': ('A401', '2070', None, 'UCAN溝通表達'),  # 6070: UCAN溝通表達
    '6080': ('A401', '2080', None, 'UCAN資訊科技'),  # 6080: UCAN資訊科技
    
    # STEAM (A601) 代碼映射
    '7010': ('A601', '4010', None, 'STEAM科學'),  # 7010: STEAM科學
    '7020': ('A601', '4020', None, 'STEAM科技'),  # 7020: STEAM科技
    '7030': ('A601', '4030', None, 'STEAM工程'),  # 7030: STEAM工程
    '7040': ('A601', '4040', None, 'STEAM數學'),  # 7040: STEAM數學
    '7050': ('A601', '4050', None, 'STEAM藝術'),  # 7050: STEAM藝術
}

# 標準欄位映射（CSV 格式和完整名稱格式）
FIELD_MAPPING = {
    # SDGs (A101)
    '1.消除貧窮': ('A101', '01', None),
    '2.消除飢餓': ('A101', '02', None),
    '3.健康與福祉': ('A101', '03', None),
    '4.教育品質': ('A101', '04', None),
    '5.性別平等': ('A101', '05', None),
    '6.淨水與衛生': ('A101', '06', None),
    '7.可負擔能源': ('A101', '07', None),
    '8.就業與經濟成長': ('A101', '08', None),
    '9.工業、創新基礎建設': ('A101', '09', None),
    '10.減少不平等': ('A101', '10', None),
    '11.永續城市': ('A101', '11', None),
    '12.責任消費與生產': ('A101', '12', None),
    '13.氣候行動': ('A101', '13', None),
    '14.海洋生態': ('A101', '14', None),
    '15.陸地生態': ('A101', '15', None),
    '16.和平與正義制度': ('A101', '16', None),
    '17.全球夥伴': ('A101', '17', None),
    
    # 主題 (A301)
    # 注意：資料庫中 A301 使用 1012 (跨域)，因為 1010 已被 A401 使用
    '跨域(主題:指標主題)': ('A301', '1012', None),
    '實作(主題:指標主題)': ('A301', '1020', '實作週次'),
    '行動導向(主題:指標主題)': ('A301', '1030', None),
    '氣候變遷(主題:指標主題)': ('A301', '1040', None),
    '淨零排放(主題:指標主題)': ('A301', '1050', None),
    '永續發展(主題:指標主題)': ('A301', '1060', None),
    '情緒管理(主題:指標主題)': ('A301', '1070', None),
    '人文關懷(主題:指標主題)': ('A301', '1080', '人文關懷週次'),
    '性別平等(主題:指標主題)': ('A301', '1090', None),
    '生命教育(主題:指標主題)': ('A301', '1100', None),
    '資訊安全(主題:指標主題)': ('A301', '1110', None),
    '資訊判讀(主題:指標主題)': ('A301', '1120', '媒體識讀或資訊判讀週次'),
    '媒體識讀(主題:指標主題)': ('A301', '1130', '媒體識讀或資訊判讀週次'),
    '智慧財產(主題:指標主題)': ('A301', '1140', None),
    '臺灣文學(主題:指標主題)': ('A301', '1150', None),
    '走讀台中山水遊學(主題:指標主題)': ('A301', '1160', None),
    '空間綠化設計(主題:指標主題)': ('A301', '1170', None),
    '生活環境創意美學(主題:指標主題)': ('A301', '1180', None),
    '運算思維程式設計(主題:指標主題)': ('A301', '1190', None),
    '創新創意(主題:指標主題)': ('A301', '1200', None),
    '設計思考(主題:指標主題)': ('A301', '1210', None),
    '生成式AI(主題:指標主題)': ('A301', '1220', None),
    '職涯發展(主題:指標主題)': ('A301', '1230', None),
    
    # UCAN (A401)
    # 注意：sub_theme_code 必須全局唯一，使用 2010-2080 範圍
    '資訊科技(UCAN)': ('A401', '1010', '資訊科技週次'),  # 已存在
    'UCAN創新': ('A401', '2010', None),
    'UCAN問題解決': ('A401', '2020', None),
    'UCAN人際互動': ('A401', '2030', None),
    'UCAN工作責任及紀律': ('A401', '2040', None),
    'UCAN持續學習': ('A401', '2050', None),
    'UCAN團隊合作': ('A401', '2060', None),
    'UCAN溝通表達': ('A401', '2070', None),
    'UCAN資訊科技應用': ('A401', '2080', None),
    
    # USR (A501)
    # 注意：sub_theme_code 必須全局唯一，使用 3010-3070 範圍
    '在地關懷(USR)': ('A501', '3010', '在地關懷週次'),
    '文化永續(USR)': ('A501', '3020', None),
    '健康促進(USR)': ('A501', '3030', None),
    '永續環境(USR)': ('A501', '3040', None),
    '產業鏈結與經濟永續(USR)': ('A501', '3050', None),
    '食品安全(USR)': ('A501', '3060', None),
    '社會實踐(USR)': ('A501', '3070', None),
    
    # STEAM (A601)
    # 注意：sub_theme_code 必須全局唯一，使用 4010-4050 範圍
    'STEAM科學': ('A601', '4010', None),
    'STEAM科技': ('A601', '4020', None),
    'STEAM工程': ('A601', '4030', None),
    'STEAM數學': ('A601', '4040', None),
    'STEAM藝術': ('A601', '4050', None),
}

# 主題定義
THEMES = {
    'A101': {
        'theme_name': '聯合國全球永續發展目標',
        'theme_short_name': 'SDGs',
        'theme_english_name': 'SDGs',
        'chinese_link': 'https://globalgoals.tw/',
        'english_link': None,
    },
    'A301': {
        'theme_name': '主導項目',
        'theme_short_name': '主導',
        'theme_english_name': 'Theme Project',
        'chinese_link': None,
        'english_link': None,
    },
    'A401': {
        'theme_name': 'UCAN',
        'theme_short_name': 'UCAN',
        'theme_english_name': 'UCAN',
        'chinese_link': None,
        'english_link': None,
    },
    'A501': {
        'theme_name': '大學社會責任實踐(USR)',
        'theme_short_name': 'USR',
        'theme_english_name': 'University Social Responsibility USR',
        'chinese_link': 'https://usr.moe.gov.tw/about/us',
        'english_link': None,
    },
    'A601': {
        'theme_name': 'STEAM',
        'theme_short_name': 'STEAM',
        'theme_english_name': 'STEAM',
        'chinese_link': None,
        'english_link': None,
    },
}

# SDGs 細項主題定義
SDGS_SUB_THEMES = {
    '01': {'name': '消除貧窮', 'english_name': 'No Poverty'},
    '02': {'name': '消除飢餓', 'english_name': 'Zero Hunger'},
    '03': {'name': '健康與福祉', 'english_name': 'Good Health and Well-being'},
    '04': {'name': '優質教育', 'english_name': 'Quality Education'},
    '05': {'name': '性別平等', 'english_name': 'Gender Equality'},
    '06': {'name': '淨水與衛生', 'english_name': 'Clean Water and Sanitation'},
    '07': {'name': '可負擔的潔淨能源', 'english_name': 'Affordable and Clean Energy'},
    '08': {'name': '尊嚴就業與經濟發展', 'english_name': 'Decent Work and Economic Growth'},
    '09': {'name': '產業創新與基礎設施', 'english_name': 'Industry, Innovation and Infrastructure'},
    '10': {'name': '減少不平等', 'english_name': 'Reduced Inequalities'},
    '11': {'name': '永續城市與社區', 'english_name': 'Sustainable Cities and Communities'},
    '12': {'name': '負責任的消費與生產', 'english_name': 'Responsible Consumption and Production'},
    '13': {'name': '氣候行動', 'english_name': 'Climate Action'},
    '14': {'name': '保育海洋生態', 'english_name': 'Life Below Water'},
    '15': {'name': '保育陸域生態', 'english_name': 'Life on Land'},
    '16': {'name': '和平正義與健全制度', 'english_name': 'Peace, Justice and Strong Institutions'},
    '17': {'name': '全球夥伴', 'english_name': 'Partnerships for the Goals'},
}

# 主題細項主題定義（A301）
# 注意：資料庫中已有 1012 (跨域)，由於 1010 已被 A401 使用，A301 使用 1012
THEME_SUB_THEMES = {
    '1012': {'name': '跨域', 'english_name': 'Interdisciplinary'},  # 使用 1012 因為 1010 被 A401 使用
    '1020': {'name': '實作', 'english_name': 'Practice'},
    '1030': {'name': '行動導向', 'english_name': 'Action-Oriented'},
    '1040': {'name': '氣候變遷', 'english_name': 'Climate Change'},
    '1050': {'name': '淨零排放', 'english_name': 'Net Zero Emissions'},
    '1060': {'name': '永續發展', 'english_name': 'Sustainable Development'},
    '1070': {'name': '情緒管理', 'english_name': 'Emotion Management'},
    '1080': {'name': '人文關懷', 'english_name': 'Humanistic Care'},
    '1090': {'name': '性別平等', 'english_name': 'Gender Equality'},
    '1100': {'name': '生命教育', 'english_name': 'Life Education'},
    '1110': {'name': '資訊安全', 'english_name': 'Information Security'},
    '1120': {'name': '資訊判讀', 'english_name': 'Information Literacy'},
    '1130': {'name': '媒體識讀', 'english_name': 'Media Literacy'},
    '1140': {'name': '智慧財產', 'english_name': 'Intellectual Property'},
    '1150': {'name': '臺灣文學', 'english_name': 'Taiwan Literature'},
    '1160': {'name': '走讀台中山水遊學', 'english_name': 'Taichung Landscape Study Tour'},
    '1170': {'name': '空間綠化設計', 'english_name': 'Space Greening Design'},
    '1180': {'name': '生活環境創意美學', 'english_name': 'Living Environment Creative Aesthetics'},
    '1190': {'name': '運算思維程式設計', 'english_name': 'Computational Thinking Programming'},
    '1200': {'name': '創新創意', 'english_name': 'Innovation and Creativity'},
    '1210': {'name': '設計思考', 'english_name': 'Design Thinking'},
    '1220': {'name': '生成式AI', 'english_name': 'Generative AI'},
    '1230': {'name': '職涯發展', 'english_name': 'Career Development'},
}

# UCAN 細項主題定義（A401）
# 注意：sub_theme_code 必須全局唯一
UCAN_SUB_THEMES = {
    '1010': {'name': '資訊科技', 'english_name': 'Information Technology'},  # 已存在
    '2010': {'name': '創新', 'english_name': 'Innovation'},  # 使用新 code
    '2020': {'name': '問題解決', 'english_name': 'Problem Solving'},  # 使用新 code
    '2030': {'name': '人際互動', 'english_name': 'Interpersonal Interaction'},  # 使用新 code
    '2040': {'name': '工作責任及紀律', 'english_name': 'Work Responsibility and Discipline'},  # 使用新 code
    '2050': {'name': '持續學習', 'english_name': 'Continuous Learning'},  # 使用新 code
    '2060': {'name': '團隊合作', 'english_name': 'Teamwork'},  # 使用新 code
    '2070': {'name': '溝通表達', 'english_name': 'Communication'},  # 使用新 code
    '2080': {'name': '資訊科技應用', 'english_name': 'Information Technology Application'},  # 使用新 code
}

# USR 細項主題定義（A501）
# 注意：sub_theme_code 必須全局唯一，使用 3010-3070 範圍
USR_SUB_THEMES = {
    '3010': {'name': '在地關懷', 'english_name': 'Local Care'},  # 使用新 code
    '3020': {'name': '文化永續', 'english_name': 'Cultural Sustainability'},  # 使用新 code
    '3030': {'name': '健康促進', 'english_name': 'Health Promotion'},  # 使用新 code
    '3040': {'name': '永續環境', 'english_name': 'Sustainable Environment'},  # 使用新 code
    '3050': {'name': '產業鏈結與經濟永續', 'english_name': 'Industry Linkage and Economic Sustainability'},  # 使用新 code
    '3060': {'name': '食品安全', 'english_name': 'Food Safety'},  # 使用新 code
    '3070': {'name': '社會實踐', 'english_name': 'Social Practice'},  # 使用新 code
}

# STEAM 細項主題定義（A601）
# 注意：sub_theme_code 必須全局唯一，使用 4010-4050 範圍
STEAM_SUB_THEMES = {
    '4010': {'name': '科學', 'english_name': 'Science'},  # 使用新 code
    '4020': {'name': '科技', 'english_name': 'Technology'},  # 使用新 code
    '4030': {'name': '工程', 'english_name': 'Engineering'},  # 使用新 code
    '4040': {'name': '數學', 'english_name': 'Mathematics'},  # 使用新 code
    '4050': {'name': '藝術', 'english_name': 'Arts'},  # 使用新 code
}

# 所有細項主題定義的映射
ALL_SUB_THEMES = {
    'A101': SDGS_SUB_THEMES,
    'A301': THEME_SUB_THEMES,
    'A401': UCAN_SUB_THEMES,
    'A501': USR_SUB_THEMES,
    'A601': STEAM_SUB_THEMES,
}


def parse_academic_year_term(year_term_str: str) -> Tuple[int, int]:
    """解析學年期字串（例如 '1141' -> 學年 114, 學期 1）"""
    if not year_term_str or len(str(year_term_str).strip()) != 4:
        raise ValueError(f"無效的學年期格式: {year_term_str}")
    year_term_str = str(year_term_str).strip()
    academic_year = int(year_term_str[:3])
    academic_term = int(year_term_str[3])
    return academic_year, academic_term


def parse_week_numbers(week_str: Optional[str]) -> Optional[List[int]]:
    """解析週次字串（例如 '2,3,4,5' -> [2, 3, 4, 5]）"""
    if not week_str or not str(week_str).strip():
        return None
    try:
        weeks = [int(w.strip()) for w in str(week_str).split(',') if w.strip()]
        return weeks if weeks else None
    except (ValueError, AttributeError):
        return None


def build_field_mapping_from_headers(headers: List[str]) -> Dict[str, tuple]:
    """根據 Excel 標題欄位動態建立欄位映射"""
    field_mapping = {}
    
    # 先加入標準 CSV 格式的映射
    standard_mapping = {
        # SDGs (A101) - 標準格式
        '1.消除貧窮': ('A101', '01', None),
        '2.消除飢餓': ('A101', '02', None),
        '3.健康與福祉': ('A101', '03', None),
        '4.教育品質': ('A101', '04', None),
        '5.性別平等': ('A101', '05', None),
        '6.淨水與衛生': ('A101', '06', None),
        '7.可負擔能源': ('A101', '07', None),
        '8.就業與經濟成長': ('A101', '08', None),
        '9.工業、創新基礎建設': ('A101', '09', None),
        '10.減少不平等': ('A101', '10', None),
        '11.永續城市': ('A101', '11', None),
        '12.責任消費與生產': ('A101', '12', None),
        '13.氣候行動': ('A101', '13', None),
        '14.海洋生態': ('A101', '14', None),
        '15.陸地生態': ('A101', '15', None),
        '16.和平與正義制度': ('A101', '16', None),
        '17.全球夥伴': ('A101', '17', None),
    }
    field_mapping.update(standard_mapping)
    
    # 處理每個標題欄位，檢查是否匹配數字代碼格式
    for header in headers:
        if not header or not str(header).strip():
            continue
        
        header_str = str(header).strip()
        
        # 檢查是否為 SDG 欄位（格式可能是 "01. 消除貧窮" 或 "01"）
        for sdg_num in range(1, 18):
            sdg_key = f'{sdg_num:02d}'
            if header_str.startswith(sdg_key):
                field_mapping[header_str] = ('A101', sdg_key, None)
                break
        
        # 檢查是否為數字代碼格式（如 "2030: 跨域" 或 "2030"）
        # 提取數字部分
        match = re.match(r'^(\d+)', header_str)
        if match:
            code = match.group(1)
            if code in EXCEL_CODE_TO_SUB_THEME:
                theme_code, sub_theme_code, week_column, _ = EXCEL_CODE_TO_SUB_THEME[code]
                field_mapping[header_str] = (theme_code, sub_theme_code, week_column)
                continue
        
        # 檢查標準格式的主題欄位
        standard_theme_fields = {
            '跨域(主題:指標主題)': ('A301', '1012', None),
            '實作(主題:指標主題)': ('A301', '1020', '實作週次'),
            '行動導向(主題:指標主題)': ('A301', '1030', None),
            '氣候變遷(主題:指標主題)': ('A301', '1040', None),
            '淨零排放(主題:指標主題)': ('A301', '1050', None),
            '永續發展(主題:指標主題)': ('A301', '1060', None),
            '情緒管理(主題:指標主題)': ('A301', '1070', None),
            '人文關懷(主題:指標主題)': ('A301', '1080', '人文關懷週次'),
            '性別平等(主題:指標主題)': ('A301', '1090', None),
            '生命教育(主題:指標主題)': ('A301', '1100', None),
            '資訊安全(主題:指標主題)': ('A301', '1110', None),
            '資訊判讀(主題:指標主題)': ('A301', '1120', '媒體識讀或資訊判讀週次'),
            '媒體識讀(主題:指標主題)': ('A301', '1130', '媒體識讀或資訊判讀週次'),
            '智慧財產(主題:指標主題)': ('A301', '1140', None),
            '臺灣文學(主題:指標主題)': ('A301', '1150', None),
            '走讀台中山水遊學(主題:指標主題)': ('A301', '1160', None),
            '空間綠化設計(主題:指標主題)': ('A301', '1170', None),
            '生活環境創意美學(主題:指標主題)': ('A301', '1180', None),
            '運算思維程式設計(主題:指標主題)': ('A301', '1190', None),
            '創新創意(主題:指標主題)': ('A301', '1200', None),
            '設計思考(主題:指標主題)': ('A301', '1210', None),
            '生成式AI(主題:指標主題)': ('A301', '1220', None),
            '職涯發展(主題:指標主題)': ('A301', '1230', None),
            '資訊科技(UCAN)': ('A401', '1010', '資訊科技週次'),
            'UCAN創新': ('A401', '2010', None),
            'UCAN問題解決': ('A401', '2020', None),
            'UCAN人際互動': ('A401', '2030', None),
            'UCAN工作責任及紀律': ('A401', '2040', None),
            'UCAN持續學習': ('A401', '2050', None),
            'UCAN團隊合作': ('A401', '2060', None),
            'UCAN溝通表達': ('A401', '2070', None),
            'UCAN資訊科技應用': ('A401', '2080', None),
            '在地關懷(USR)': ('A501', '3010', '在地關懷週次'),
            '文化永續(USR)': ('A501', '3020', None),
            '健康促進(USR)': ('A501', '3030', None),
            '永續環境(USR)': ('A501', '3040', None),
            '產業鏈結與經濟永續(USR)': ('A501', '3050', None),
            '食品安全(USR)': ('A501', '3060', None),
            '社會實踐(USR)': ('A501', '3070', None),
            'STEAM科學': ('A601', '4010', None),
            'STEAM科技': ('A601', '4020', None),
            'STEAM工程': ('A601', '4030', None),
            'STEAM數學': ('A601', '4040', None),
            'STEAM藝術': ('A601', '4050', None),
        }
        
        if header_str in standard_theme_fields:
            field_mapping[header_str] = standard_theme_fields[header_str]
    
    return field_mapping


def read_excel_file(file_path: str) -> Tuple[List[Dict], Dict[str, tuple]]:
    """讀取 Excel 檔案並返回資料和動態建立的欄位映射"""
    print("\n" + "=" * 60)
    print("階段 4: 讀取 Excel 檔案")
    print("=" * 60)
    
    rows = []
    headers = []
    
    if HAS_OPENPYXL:
        # 使用 openpyxl 讀取
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 讀取第一行作為標題
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            print(f"標題欄位數: {len(headers)}")
            print(f"前 10 個標題: {headers[:10]}")
            
            # 動態建立欄位映射
            field_mapping = build_field_mapping_from_headers(headers)
            print(f"已建立 {len(field_mapping)} 個欄位映射")
            
            # 讀取資料行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        # 處理數值類型的 cell
                        if cell.value is None:
                            row_dict[header] = ''
                        elif isinstance(cell.value, (int, float)):
                            # 如果是數字，轉為字串，但保留整數的格式
                            if isinstance(cell.value, float) and cell.value.is_integer():
                                row_dict[header] = str(int(cell.value))
                            else:
                                row_dict[header] = str(cell.value)
                        else:
                            row_dict[header] = str(cell.value).strip() if cell.value else ''
                rows.append(row_dict)
            
            print(f"讀取到 {len(rows)} 筆記錄")
            return rows, field_mapping
        except Exception as e:
            print(f"使用 openpyxl 讀取失敗: {e}")
            print("嘗試其他方法...")
            import traceback
            traceback.print_exc()
    
    # 如果沒有 openpyxl 或讀取失敗，嘗試使用 pandas
    try:
        import pandas as pd
        df = pd.read_excel(file_path)
        headers = list(df.columns)
        rows = df.to_dict('records')
        # 轉換所有值為字串，處理 NaN
        for row in rows:
            for key, value in row.items():
                if pd.isna(value):
                    row[key] = ''
                else:
                    row[key] = str(value).strip()
        
        # 動態建立欄位映射
        field_mapping = build_field_mapping_from_headers(headers)
        
        print(f"使用 pandas 讀取到 {len(rows)} 筆記錄")
        print(f"已建立 {len(field_mapping)} 個欄位映射")
        return rows, field_mapping
    except ImportError:
        print("錯誤: 需要安裝 openpyxl 或 pandas 來讀取 Excel 檔案")
        print("請執行: pip install openpyxl 或 pip install pandas")
        sys.exit(1)
    except Exception as e:
        print(f"讀取 Excel 檔案失敗: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


# 從 import_course_data_from_csv.py 複製其他函數
async def check_and_create_themes(conn) -> Dict[str, int]:
    """檢查並建立缺失的主題"""
    print("\n" + "=" * 60)
    print("階段 1: 檢查並建立主題")
    print("=" * 60)
    
    # 查詢現有主題
    existing_themes = await ThemeDAO.get_all_themes(conn)
    existing_theme_codes = {t['theme_code'] for t in existing_themes}
    
    print(f"現有主題數: {len(existing_themes)}")
    for theme in existing_themes:
        print(f"  - {theme['theme_code']}: {theme['theme_name']}")
    
    # 建立缺失的主題
    created_count = 0
    for theme_code, theme_info in THEMES.items():
        if theme_code not in existing_theme_codes:
            try:
                await ThemeDAO.create_theme(
                    conn,
                    theme_code=theme_code,
                    theme_name=theme_info['theme_name'],
                    theme_short_name=theme_info['theme_short_name'],
                    theme_english_name=theme_info['theme_english_name'],
                    chinese_link=theme_info.get('chinese_link'),
                    english_link=theme_info.get('english_link'),
                    created_by='excel_import'
                )
                print(f"  ✓ 建立主題: {theme_code} - {theme_info['theme_name']}")
                created_count += 1
            except Exception as e:
                error_str = str(e).lower()
                if 'unique constraint' in error_str or 'ora-00001' in error_str:
                    print(f"  ℹ️  主題已存在: {theme_code}")
                else:
                    print(f"  ❌ 建立主題失敗 {theme_code}: {e}")
    
    print(f"\n新建主題數: {created_count}")
    return {'created': created_count, 'existing': len(existing_themes)}


async def check_and_create_sub_themes(conn) -> Dict[str, int]:
    """檢查並建立缺失的細項主題"""
    print("\n" + "=" * 60)
    print("階段 2: 檢查並建立細項主題")
    print("=" * 60)
    
    # 查詢現有細項主題
    existing_sub_themes = await SubThemeDAO.get_all_sub_themes(conn)
    existing_sub_theme_keys = {(st['theme_code'], st['sub_theme_code']) for st in existing_sub_themes}
    
    print(f"現有細項主題數: {len(existing_sub_themes)}")
    
    # 建立缺失的細項主題
    created_count = 0
    failed_count = 0
    for theme_code, sub_themes_dict in ALL_SUB_THEMES.items():
        theme_missing = []
        for sub_theme_code, sub_theme_info in sub_themes_dict.items():
            key = (theme_code, sub_theme_code)
            if key not in existing_sub_theme_keys:
                theme_missing.append((sub_theme_code, sub_theme_info))
        
        if theme_missing:
            # 先獲取主題ID
            theme = await ThemeDAO.get_theme_by_code(conn, theme_code)
            if not theme:
                print(f"  ❌ 主題 {theme_code} 不存在，跳過")
                continue
            
            print(f"\n  處理主題 {theme_code}，需要建立 {len(theme_missing)} 個細項主題:")
            for sub_theme_code, sub_theme_info in theme_missing:
                try:
                    await SubThemeDAO.create_sub_theme(
                        conn,
                        coures_themes_id=theme['id'],
                        sub_theme_code=sub_theme_code,
                        sub_theme_name=sub_theme_info['name'],
                        sub_theme_english_name=sub_theme_info['english_name'],
                        created_by='excel_import'
                    )
                    print(f"    ✓ 建立: {sub_theme_code} - {sub_theme_info['name']}")
                    created_count += 1
                except Exception as e:
                    error_str = str(e).lower()
                    if 'unique constraint' in error_str or 'ora-00001' in error_str:
                        print(f"    ℹ️  已存在: {sub_theme_code} - {sub_theme_info['name']}")
                    else:
                        print(f"    ❌ 失敗: {sub_theme_code} - {sub_theme_info['name']}: {e}")
                        failed_count += 1
    
    print(f"\n新建細項主題數: {created_count}")
    return {'created': created_count, 'existing': len(existing_sub_themes)}


async def create_school_year_settings(conn, academic_years_terms: Set[Tuple[int, int]]) -> Dict[str, int]:
    """建立學年期主題設定和細項主題設定"""
    print("\n" + "=" * 60)
    print("階段 3: 建立學年期設定")
    print("=" * 60)
    
    theme_settings_created = 0
    sub_theme_settings_created = 0
    
    # 主題設定配置（根據實際需求調整）
    theme_settings_config = {
        'A101': {'fill_in_week_enabled': False, 'scale_max': 5},
        'A301': {'fill_in_week_enabled': True, 'scale_max': 5},
        'A401': {'fill_in_week_enabled': True, 'scale_max': 5},
        'A501': {'fill_in_week_enabled': True, 'scale_max': 5},
        'A601': {'fill_in_week_enabled': False, 'scale_max': 5},
    }
    
    for academic_year, academic_term in sorted(academic_years_terms):
        print(f"\n處理學年期: {academic_year}-{academic_term}")
        
        for theme_code in THEMES.keys():
            # 檢查主題設定是否已存在
            existing_setting = await SchoolYearThemeSettingsDAO.get_school_year_theme_setting(
                conn, academic_year, academic_term, theme_code
            )
            
            if not existing_setting:
                try:
                    config = theme_settings_config.get(theme_code, {
                        'fill_in_week_enabled': False,
                        'scale_max': 5
                    })
                    
                    await SchoolYearThemeSettingsDAO.create_school_year_theme_setting(
                        conn,
                        academic_year=academic_year,
                        academic_term=academic_term,
                        theme_code=theme_code,
                        fill_in_week_enabled=config['fill_in_week_enabled'],
                        scale_max=config['scale_max'],
                        select_most_relevant_sub_theme_enabled=False,
                        created_by='excel_import'
                    )
                    # 計算該主題的細項主題數量（由主題設定自動建立）
                    sub_themes = await SubThemeDAO.get_sub_themes_by_theme_code(conn, theme_code)
                    sub_theme_settings_created += len(sub_themes)
                    print(f"  ✓ 建立主題設定: {theme_code} (包含 {len(sub_themes)} 個細項主題設定)")
                    theme_settings_created += 1
                except Exception as e:
                    error_str = str(e).lower()
                    if 'unique constraint' in error_str or 'ora-00001' in error_str:
                        print(f"  ℹ️  主題設定已存在: {theme_code}")
                    else:
                        print(f"  ❌ 建立主題設定失敗 {theme_code}: {e}")
            else:
                print(f"  ℹ️  主題設定已存在: {theme_code}")
                # 檢查並建立缺失的細項主題設定
                try:
                    # 取得該主題的所有細項主題
                    all_sub_themes = await SubThemeDAO.get_sub_themes_by_theme_code(conn, theme_code)
                    print(f"    → 該主題共有 {len(all_sub_themes)} 個細項主題")
                    
                    # 取得已存在的細項主題設定
                    existing_sub_theme_settings = await SchoolYearSubThemeSettingsDAO.get_school_year_sub_theme_settings_by_year_and_theme(
                        conn, academic_year, academic_term, theme_code
                    )
                    # 只計算真正存在的細項主題設定（created_at 不為 None 表示設定存在）
                    existing_sub_theme_codes = {
                        st['sub_theme_code'] for st in existing_sub_theme_settings 
                        if st.get('created_at') is not None
                    }
                    print(f"    → 已有 {len(existing_sub_theme_codes)} 個細項主題設定（共 {len(existing_sub_theme_settings)} 個細項主題）")
                    
                    # 找出缺失的細項主題設定
                    missing_sub_themes = [st for st in all_sub_themes if st['sub_theme_code'] not in existing_sub_theme_codes]
                    
                    if missing_sub_themes:
                        print(f"    → 需要建立 {len(missing_sub_themes)} 個缺失的細項主題設定")
                        # 建立缺失的細項主題設定
                        missing_count = 0
                        for sub_theme in missing_sub_themes:
                            try:
                                await SchoolYearSubThemeSettingsDAO.create_school_year_sub_theme_setting(
                                    conn,
                                    academic_year=academic_year,
                                    academic_term=academic_term,
                                    theme_code=theme_code,
                                    sub_theme_code=sub_theme['sub_theme_code'],
                                    enabled=True,
                                    created_by='excel_import'
                                )
                                missing_count += 1
                                print(f"      ✓ 建立: {sub_theme['sub_theme_code']} - {sub_theme['sub_theme_name']}")
                            except Exception as e:
                                error_str = str(e).lower()
                                if 'unique constraint' in error_str or 'ora-00001' in error_str:
                                    print(f"      ℹ️  已存在: {sub_theme['sub_theme_code']} - {sub_theme['sub_theme_name']}")
                                else:
                                    print(f"      ❌ 失敗: {sub_theme['sub_theme_code']} - {sub_theme['sub_theme_name']}: {e}")
                        
                        if missing_count > 0:
                            sub_theme_settings_created += missing_count
                            print(f"    → 成功建立 {missing_count} 個細項主題設定")
                    else:
                        print(f"    → 所有細項主題設定已存在")
                except Exception as e:
                    print(f"    ⚠️  檢查細項主題設定時發生錯誤: {e}")
                    import traceback
                    traceback.print_exc()
    
    print(f"\n新建主題設定數: {theme_settings_created}")
    print(f"新建細項主題設定數: {sub_theme_settings_created} (由主題設定自動建立)")
    
    return {
        'theme_settings_created': theme_settings_created,
        'sub_theme_settings_created': sub_theme_settings_created
    }


async def import_course_entries(conn, excel_rows: List[Dict], field_mapping: Dict[str, tuple]) -> Dict[str, int]:
    """匯入課程填寫記錄"""
    print("\n" + "=" * 60)
    print("階段 5: 匯入課程填寫記錄")
    print("=" * 60)
    
    success_count = 0
    skip_count = 0
    error_count = 0
    errors = []
    
    total_records = 0
    
    for row_idx, row in enumerate(excel_rows, 1):
        # 解析學年期
        try:
            academic_year, academic_term = parse_academic_year_term(row.get('學年期', ''))
        except Exception as e:
            error_count += 1
            errors.append(f"第 {row_idx} 行: 學年期解析失敗 - {e}")
            continue
        
        subj_no = str(row.get('OPMS_COURSE_NO', '')).strip()
        ps_class_nbr = str(row.get('PS_CLASS_NBR', '')).strip()
        
        if not subj_no or not ps_class_nbr:
            error_count += 1
            errors.append(f"第 {row_idx} 行: 缺少必要欄位 OPMS_COURSE_NO 或 PS_CLASS_NBR")
            continue
        
        # 處理每個有值的欄位
        for excel_column, (theme_code, sub_theme_code, week_column) in field_mapping.items():
            if excel_column not in row:
                continue
            
            indicator_value = str(row[excel_column]).strip() if row.get(excel_column) else None
            if not indicator_value or indicator_value == '' or indicator_value == 'None' or indicator_value.lower() == 'nan':
                continue
            
            # 取得週次資料
            week_numbers = None
            if week_column and week_column in row:
                week_numbers = parse_week_numbers(row.get(week_column))
            
            # 檢查記錄是否已存在
            existing_entry = await CourseEntriesDAO.get_course_entry(
                conn, subj_no, ps_class_nbr, academic_year, academic_term, sub_theme_code
            )
            
            if existing_entry:
                skip_count += 1
                continue
            
            # 建立記錄
            try:
                await CourseEntriesDAO.create_course_entry(
                    conn,
                    subj_no=subj_no,
                    ps_class_nbr=ps_class_nbr,
                    academic_year=academic_year,
                    academic_term=academic_term,
                    sub_theme_code=sub_theme_code,
                    indicator_value=str(indicator_value),
                    week_numbers=week_numbers,
                    is_most_relevant=False,
                    created_by='excel_import'
                )
                success_count += 1
                total_records += 1
                
                if total_records % 100 == 0:
                    print(f"  已處理 {total_records} 筆記錄...")
                    
            except Exception as e:
                error_count += 1
                error_msg = f"第 {row_idx} 行, 課程 {subj_no}-{ps_class_nbr}, 細項主題 {sub_theme_code}: {e}"
                errors.append(error_msg)
                if len(errors) <= 10:  # 只記錄前 10 個錯誤
                    print(f"  ❌ {error_msg}")
    
    print(f"\n匯入結果:")
    print(f"  ✓ 成功匯入: {success_count} 筆")
    print(f"  ℹ️  跳過 (已存在): {skip_count} 筆")
    print(f"  ❌ 錯誤: {error_count} 筆")
    
    if errors and len(errors) > 10:
        print(f"\n前 10 個錯誤:")
        for error in errors[:10]:
            print(f"  - {error}")
    elif errors:
        print(f"\n錯誤詳情:")
        for error in errors:
            print(f"  - {error}")
    
    return {
        'success': success_count,
        'skip': skip_count,
        'error': error_count,
        'errors': errors
    }


async def main():
    """主函數"""
    excel_file_path = '塞入的資料.xlsx'
    
    print("=" * 60)
    print("Excel 課程資料匯入工具")
    print("=" * 60)
    
    # 連線資料庫
    dsn = get_database_dsn()
    conn = None
    try:
        print(f"\n連線資料庫: {dsn}")
        conn = await asyncio.to_thread(
            oracledb.connect,
            user=settings.db_user,
            password=settings.db_password,
            dsn=dsn
        )
        print("✓ 資料庫連線成功")
        
        # 階段 1: 檢查並建立主題
        theme_stats = await check_and_create_themes(conn)
        
        # 階段 2: 檢查並建立細項主題
        sub_theme_stats = await check_and_create_sub_themes(conn)
        
        # 階段 3: 讀取 Excel 並分析需要的學年期
        excel_rows, field_mapping = read_excel_file(excel_file_path)
        academic_years_terms = set()
        for row in excel_rows:
            try:
                academic_year, academic_term = parse_academic_year_term(row.get('學年期', ''))
                academic_years_terms.add((academic_year, academic_term))
            except:
                pass
        
        # 階段 4: 建立學年期設定
        settings_stats = await create_school_year_settings(conn, academic_years_terms)
        
        # 階段 5: 匯入課程記錄
        import_stats = await import_course_entries(conn, excel_rows, field_mapping)
        
        # 提交所有變更
        await asyncio.to_thread(conn.commit)
        print("\n✓ 所有變更已提交")
        
        # 輸出統計報告
        print("\n" + "=" * 60)
        print("匯入統計報告")
        print("=" * 60)
        print(f"主題:")
        print(f"  - 現有: {theme_stats['existing']}")
        print(f"  - 新建: {theme_stats['created']}")
        print(f"細項主題:")
        print(f"  - 現有: {sub_theme_stats['existing']}")
        print(f"  - 新建: {sub_theme_stats['created']}")
        print(f"學年期設定:")
        print(f"  - 主題設定新建: {settings_stats['theme_settings_created']}")
        print(f"  - 細項主題設定新建: {settings_stats['sub_theme_settings_created']}")
        print(f"課程記錄:")
        print(f"  - 成功匯入: {import_stats['success']}")
        print(f"  - 跳過: {import_stats['skip']}")
        print(f"  - 錯誤: {import_stats['error']}")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n❌ 發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            await asyncio.to_thread(conn.rollback)
        sys.exit(1)
    finally:
        if conn:
            await asyncio.to_thread(conn.close)
            print("\n資料庫連線已關閉")


if __name__ == "__main__":
    asyncio.run(main())

